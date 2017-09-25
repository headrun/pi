import smtplib
from table_schemas.generic_functions import *
from table_schemas.pi_db_operations import sender_mail_pi, sender_pwd_pi
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from terminal_commands import *
import glob
import openpyxl as px


class Piparsing(object):

	def parameters(self):
		con, cur = get_mysql_connection(DB_HOST, DB_NAME_REQ, '')
		print DB_NAME_REQ
                current_path = os.path.dirname(os.path.abspath(__file__))
		present_date = re.sub('\.(.*)','', str(datetime.datetime.now()))
                social_processing_path = os.path.join(
		current_path, 'excelfiles')
		return con, cur, current_path, social_processing_path, present_date

	def main(self, email_from_list):
		con, cur, current_path, social_processing_path, present_date = self.parameters()
		files_list = glob.glob(social_processing_path+'/*.xlsx')
		if not files_list:
			files_list = glob.glob(social_processing_path+'/*.xlsx*')
		twitter_check, linkedin_check, facebook_check = ['']*3
	        if files_list:
			final_send_mail = []
			for _file in files_list:
				ws_ = px.load_workbook(_file, use_iterators=True)
				sheet_list = ws_.get_sheet_names()
				if not sheet_list:
					continue
				for xl_ in sheet_list[0:1]:
					sheet_ = ws_.get_sheet_by_name(name=xl_)
					row_check = 0
					email_address, linkedin_profile, ida,\
					firstname, lastname, key_,\
					twitter_profile, facebook_profile = ['']*8
					for row in sheet_.iter_rows():
						if row_check > 0:
							row_check += 1
							continue
						counter = 0
						for i in row:
							if not i.value: continue
							lower_head = i.value.lower()
							if 'key' in lower_head:
								key_ = counter
							elif 'linkedin'  in lower_head:
								linkedin_profile = counter
							elif 'email' in lower_head:
								email_address = counter
							elif 'id' in lower_head and 'linkedin' not in lower_head and 'email' not in lower_head and 'linkdin_url' not in lower_head:
								ida = counter
							elif 'first name' in lower_head:
								firstname = counter
							elif 'last name' in lower_head:
								lastname = counter
							elif 'twitter' in lower_head:
								twitter_profile = counter
							elif 'facebook' in lower_head:
								facebook_profile = counter
							elif 's.no' in lower_head:
								snoa = counter
							counter += 1
						break
					m = 0
					final_indents = []
					_indents = {}
					for row in sheet_.iter_rows():
						temp_rows = []
						for i in row:
							temp_rows.append(i.value)
						final_indents.append(temp_rows)
					empty_dic = {}
					#present_date = re.sub('\.(.*)','', str(datetime.datetime.now()))
					for row in final_indents:
						row = map(lambda x_inner: '' if x_inner == None else x_inner, row)
						row_check = range(len(row))
						meta_date_from_browse, sk = {}, ''
						if email_address != '' and \
						email_address in row_check:
							email_addressf = normalize(
							str(row[email_address].encode('utf-8')))
							meta_date_from_browse.update(
						{"email_address": email_addressf.strip('"')})
						if ida != '' and ida in row_check:
							idaf = normalize(str(row[ida]))
							meta_date_from_browse.update({"id":idaf})
						if firstname != '' and firstname in row_check:
							firstnamef = normalize(str(row[firstname].encode('utf-8')))
							meta_date_from_browse.update({"firstname":firstnamef})
						if lastname != '' and lastname in row_check:
							lastnamef = normalize(str(row[lastname].encode('utf-8')))
							meta_date_from_browse.update({"lastname":lastnamef})
						if key_ != '' and key_ in row_check:
							key_f = normalize(str(row[key_].encode('utf-8')))
							meta_date_from_browse.update({"keys":key_f})
						if snoa != '' and snoa in row_check:
							snoaf = normalize(str(row[snoa]))
							meta_date_from_browse.update({"sno":snoaf})
                                                if linkedin_profile != '' and linkedin_profile in row_check:
                                                        linkedin_profilef = normalize(str(row[linkedin_profile].encode('utf-8')))
                                                        meta_date_from_browse.update({"linkedin_url":linkedin_profilef})
                                                        sk = md5(normalize(linkedin_profilef))
                                                        linkedin_profilef = self.linkedin_pub_url(linkedin_profilef)
							if linkedin_profilef and 'linkedin.com' in linkedin_profilef:
								linkedin_profilef = MySQLdb.escape_string(linkedin_profilef)
								values = ('linkedin_crawl', sk, linkedin_profilef, 'linkedin', 0,MySQLdb.escape_string(json.dumps(meta_date_from_browse)), MySQLdb.escape_string(json.dumps(meta_date_from_browse)))
								execute_query(cur, insert_script_query%values)
						if twitter_profile != '' and twitter_profile in row_check:
							twitter_profilef = normalize(str(row[twitter_profile].encode('utf-8')))
							twitter_profilef = MySQLdb.escape_string(twitter_profilef)
							sk = twitter_profilef.split('/')[-1].strip()
							met_browse = meta_date_from_browse.get('email_address', '')
							if twitter_profilef and 'twitter.com' in twitter_profilef:
								#values = ('twitter_crawl', sk, twitter_profilef, 'twitter', 0, MySQLdb.escape_string(met_browse), MySQLdb.escape_string(met_browse))
								values = ('twitter_crawl', sk, twitter_profilef, 'twitter', 0, MySQLdb.escape_string(json.dumps(meta_date_from_browse)), MySQLdb.escape_string(json.dumps(meta_date_from_browse)))
								execute_query(cur, insert_script_query%values)
						if facebook_profile != '' and facebook_profile in row_check:
							facebook_profilef = normalize(str(row[facebook_profile].encode('utf-8')))
							facebook_profilef = MySQLdb.escape_string(facebook_profilef)
							meta_date_from_browse.update({"mbasic_url":facebook_profilef.replace('www', 'mbasic')})
							meta_date_from_browse.pop('linkedin_url', None)
							sk = md5(normalize(facebook_profilef))
							if facebook_profilef and 'facebook.com' in facebook_profilef:
								facebook_profilef = MySQLdb.escape_string(facebook_profilef)
								values = ('facebook_crawl', sk, facebook_profilef, 'facebook', 0, MySQLdb.escape_string(json.dumps(meta_date_from_browse)), MySQLdb.escape_string(json.dumps(meta_date_from_browse)))
								execute_query(cur, insert_script_query%values)
					filenamei = "%s%s%s" % (_file.split('/')[-1].split('.')[0], ' on ', str(datetime.datetime.now().date()))
					if twitter_profile != '' and final_indents:
						len_tw_total, len_tw_dup, tw_dup_urls = \
						self.getindex_value(twitter_profile, final_indents)
						an_lentw = self.get_total_curdate('twitter_crawl', cur, present_date)
						final_send_mail.append([len_tw_total, len_tw_dup, tw_dup_urls, an_lentw, 'twitter', filenamei, xl_])
						if an_lentw and an_lentw != 0:
							twitter_check = 'on'
					if linkedin_profile != '' and final_indents:
						len_lk_total, len_lk_dup, lk_dup_urls = \
						self.getindex_value(linkedin_profile, final_indents)
						an_lenlk = self.get_total_curdate('linkedin_crawl', cur, present_date)
						final_send_mail.append([len_lk_total, len_lk_dup, lk_dup_urls, an_lenlk, 'linkedin', filenamei, xl_])
						if an_lenlk and an_lenlk != 0:
							linkedin_check = 'on'
					if facebook_profile != '' and final_indents:
						len_fb_total, len_fb_dup, fb_dup_urls = \
						self.getindex_value(facebook_profile, final_indents)
						an_lenfb = self.get_total_curdate('facebook_crawl', cur, present_date)
						final_send_mail.append([len_fb_total, len_fb_dup, fb_dup_urls, an_lenfb, 'facebook', filenamei, xl_])
						if an_lenfb and an_lenfb != 0:
							facebook_check = 'on'
				#cmd = ('"' + _file + '"')
				os.remove(_file)
			if final_send_mail:
				self.alert_mail(final_send_mail, email_from_list)
				
		close_mysql_connection(con, cur)
		if linkedin_check or facebook_check or twitter_check:
			lastseen_date = re.sub('\.(.*)','', str(datetime.datetime.now()))
			Commands().main(present_date, linkedin_check, facebook_check, twitter_check, email_from_list, lastseen_date)	
		#return present_date

	def get_total_curdate(self, table, cur, present_date):
		total_count = "select count(*) from %s where modified_at >= '%s'" % (table, present_date)
		t_count = fetchmany(cur, total_count)
		return str(t_count[0][0])

        def alert_mail(self, final_send_mail, email_from_list):
                sender_mail = sender_mail_pi
                receivers_mail_list = email_from_list
                sender, receivers  = sender_mail, ','.join(receivers_mail_list)
		che_here = ''
                msg = MIMEMultipart('alternative')
		for indexr, rows in enumerate(final_send_mail):
			let_total, len_dup, dup_urls, db_total, table_name, sheet_name, inner_shname = rows
			if indexr == 0:
				msg['Subject'] = 'Test mail: Stats for profiles in sheet %s' % (sheet_name)
				mas = '<html><head><link href="http://getbootstrap.com/dist/css/bootstrap.css" rel="stylesheet"></head>'
			if db_total:
				mas += '<h3>For %s</h3></br>' % table_name
				mas += '<p>Total : %s</p></br>' % let_total
				if dup_urls:
					mas += '<p>Duplicates : %s</p></br>' %len_dup
					dupj_urls = ',  '.join(dup_urls)
					mas += '<p> Duplicate urls : %s</br>' % dupj_urls
					mas += '<p> Exact Count : %s</p></br>' % db_total
				else:
					'<p> Duplicates: 0'
				if int(db_total) != int(let_total) and len(dup_urls)+int(db_total) != int(let_total):
					mas += '<p> Mismatch in db total count and sheet count. Check the stats once</p></br>'
					mas += '<p> db count : %s</p></br>' % db_total
					mas += '<p> total count in sheet : %s </p></br>' % let_total
				eta_minutes = ''
				if 'twitter' in table_name:
					eta_minutes = int(db_total) * 0.08
				elif 'linkedin' in table_name:
					eta_minutes = int(db_total) * 3.5
				elif 'facebook' in table_name:
					eta_minutes  = int(db_total) * 1
				if db_total not in ['0', 0]:
					mas += '<p> Exact ETA : %s minutes</p></br>' % eta_minutes
				che_here = 'yes'
		if che_here:
			msg['From'] = sender
			msg['To'] = receivers
			tem = MIMEText(''.join(mas), 'html')
			msg.attach(tem)
			s = smtplib.SMTP('smtp.gmail.com:587')
			s.ehlo()
			s.starttls()
			s.login(sender_mail, sender_pwd_pi)
			s.sendmail(sender, receivers_mail_list, msg.as_string())
			s.quit()



	def getindex_value(self, indexi, final_indents):
		total_countt_urls = []
		for fi in final_indents[1:]:
			try: total_countt_urls.append(fi[indexi])
			except:pass
		total_countt_urls = filter(None,total_countt_urls)
		duplicates_count_urls = [item for item, count in collections.Counter(total_countt_urls).items() if count > 1]
		len_total_countt_urls = len(total_countt_urls)
		len_duplicates_count_urls = len(duplicates_count_urls)
		return len_total_countt_urls, len_duplicates_count_urls, duplicates_count_urls

	def linkedin_pub_url(self, linkedin_profilef):
		if 'http:' in linkedin_profilef:
			linkedin_profilef = linkedin_profilef.replace('http:', 'https:')
	        if 'id.www' in linkedin_profilef:
        		linkedin_profilef = linkedin_profilef.replace('id.www', 'https://www')
	        if 'www.linkedin.com' and 'https:' not in linkedin_profilef:
        		linkedin_profilef = linkedin_profilef.replace('www.', 'https://www')
        	if linkedin_profilef.startswith('linkedin.com'):
            		linkedin_profilef = linkedin_profilef.replace('linkedin.com', 'https://www.linkedin.com')
	        if 'https:' not in linkedin_profilef:
        		linkedin_profilef = re.sub('(\D+)\.linkedin.com', 'https://www.linkedin.com', linkedin_profilef) 
	        linkedin_profilef = re.sub('https://(.*?).linkedin.com/', 'https://www.linkedin.com/', linkedin_profilef)
        	if linkedin_profilef.endswith('/en') or linkedin_profilef.endswith('/fr'):
            		linkedin_profilef = linkedin_profilef[:-3]
        	linkedin_profilef = linkedin_profilef.strip('"').strip().strip("'").strip().strip('/').strip()
        	if not linkedin_profilef.startswith('https://www.linkedin.com'):
            		linkedin_profilef = ''.join(re.findall('.*(https://.*)', linkedin_profilef))
        	if '/pub/' in linkedin_profilef:
            		cv = ''.join(filter(None, re.split('https://www.linkedin.com/pub/.*?/(.*)', linkedin_profilef))).split('/')[::-1]
            		cv[0] = cv[0].zfill(3)
            		cv[1] = cv[1].zfill(3)
            		if cv[-1] == '0': del cv[-1]
            		linkedin_profilef = ('%s%s%s%s' % ('https://www.linkedin.com/in/', ''.join(re.findall('https://www.linkedin.com/pub/(.*?)/.*', linkedin_profilef)), '-', ''.join(cv)))
        	return linkedin_profilef

						
if __name__ == '__main__':
    Piparsing().main(' ')
