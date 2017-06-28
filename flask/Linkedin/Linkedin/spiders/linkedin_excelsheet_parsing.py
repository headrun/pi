from linkedin_voyager_functions import *
import glob
import openpyxl as px

class Linkedinparsing(object):

	def __init__(self):
		self.con, self.cur = get_mysql_connection(DB_HOST, DB_NAME_REQ, '')
                current_path =  os.path.dirname(os.path.abspath(__file__))
                self.social_processing_path = os.path.join(current_path, 'excelfiles')
		self.query = 'insert into linkedin_crawl(sk, url, content_type ,crawl_status, meta_data,created_at, modified_at) values(%s, %s, %s, %s, %s,now(), now()) on duplicate key update modified_at=now(), content_type=%s, crawl_status=0,meta_data=%s'

	def __del__(self):
		close_mysql_connection(self.con, self.cur)

	def main(self):
		files_list = glob.glob(self.social_processing_path+'/*.xlsx')
	        if files_list:
			for _file in files_list:
				ws_ = px.load_workbook(_file, use_iterators=True)
				sheet_list = ws_.get_sheet_names()
				for xl_ in sheet_list:
					sheet_ = ws_.get_sheet_by_name(name=xl_)
					row_check = 0
					email_address = linkedin_profile = ida = firstname = lastname = key_ = 0
					for row in sheet_.iter_rows():
						if row_check > 0:
							row_check += 1
							continue
						counter = 0
						for i in row:
							if not i.value: continue
							lower_head = i.value.lower()
							if 'key' in lower_head:
								key_ = counter
							elif 'linkedin'  in lower_head:
								linkedin_profile = counter
							elif 'email' in lower_head:
								email_address = counter
							counter += 1
						break
					m = 0
					final_indents = []
					_indents = {}
					for row in sheet_.iter_rows():
						temp_rows = []
						for i in row:
							temp_rows.append(i.value)
						final_indents.append(temp_rows)
					empty_dic = {}
					counter_ = 70
					for row in final_indents:
						email_addressf = normalize(row[email_address])
						linkedin_profilef = ''
						try: linkedin_profilef = normalize(row[linkedin_profile])
						except: pass
						if not linkedin_profilef: continue
						if linkedin_profilef.strip() == 'linkedin': continue
						meta_date_from_browse = {}
						meta_date_from_browse.update({"linkedin_url":linkedin_profilef})
						meta_date_from_browse.update({"email_address":email_addressf})
						crawl_status = 0
						rows_ = []
						if linkedin_profilef.count('http') > 1: 
							inner_results = [i.replace('s://','https://').replace(', ht','').replace('LI Inmail Sent','').strip().strip(',') for i in filter(None,re.split('http*',linkedin_profilef))]
							for inr in inner_results:
								if (inr != 'https://de.linkedin.com/pub') and ('philippe-nieuwjaer') not in inr: rows_.append(inr)
						else: rows_.append(linkedin_profilef)
						for rs in rows_:
							linkedin_profilef = rs
							if 'linkedin.com' not in normalize(linkedin_profilef): crawl_status = 10
							if 'http:' in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('http:','https:')
							if 'id.www' in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('id.www','https://www')
							if 'www.linkedin.com' and 'https:' not in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('www.','https://www')
							if linkedin_profilef.startswith('linkedin.com'): linkedin_profilef = linkedin_profilef.replace('linkedin.com','https://www.linkedin.com')
							if 'https:' not in linkedin_profilef:
								linkedin_profilef = re.sub('(\D+)\.linkedin.com','https://www.linkedin.com',linkedin_profilef) 
							linkedin_profilef = re.sub('https://(.*?).linkedin.com/','https://www.linkedin.com/',linkedin_profilef)
							if linkedin_profilef.endswith('/en') or linkedin_profilef.endswith('/fr'): linkedin_profilef = linkedin_profilef[:-3]
							linkedin_profilef = linkedin_profilef.strip('"').strip().strip("'").strip().strip('/').strip()
							if not linkedin_profilef.startswith('https://www.linkedin.com') and crawl_status!=10: linkedin_profilef = ''.join(re.findall('.*(https://.*)', linkedin_profilef))
							if '/pub/' in linkedin_profilef:
								cv = ''.join(filter(None,re.split('https://www.linkedin.com/pub/.*?/(.*)',linkedin_profilef))).split('/')[::-1]
								cv[0] = cv[0].zfill(3)
								cv[1] = cv[1].zfill(3)
								if cv[-1] == '0': del cv[-1]
								linkedin_profilef = ( '%s%s%s%s'%('https://www.linkedin.com/in/',''.join(re.findall('https://www.linkedin.com/pub/(.*?)/.*',linkedin_profilef)),'-',''.join(cv)))
							if 'linkedin.com' not in linkedin_profilef: continue
							counter_ += 1
							sk = md5("%s%s"%(normalize(email_addressf),normalize( linkedin_profilef)))
							linkedin_profilef = linkedin_profilef.replace('pubwww.linkedin.comhttps:','').replace('"','')
							values = (sk, linkedin_profilef, 'linkedin', crawl_status, json.dumps(meta_date_from_browse),'linkedin', json.dumps(meta_date_from_browse))
							recof = fetchmany(self.cur, 'select * from linkedin_crawl where sk="%s"'%sk)
							if not recof:
								self.cur.execute(self.query, values)
								print 'nodup'
								print linkedin_profilef
								print '>>>>>>>'
							else:
								print values
								print '*************************'
		else:
		    for prof_url in open('linkedin_file.py'):
			if prof_url != '\n':
				prof_url = prof_url.replace('\n','')
				sk = md5.md5(prof_url).hexdigest()
				meta_date_from_browse = {}
				values = (sk, prof_url, 'linkedin', 0,json.dumps(meta_date_from_browse),'linkedin', json.dumps(meta_date_from_browse))
				self.cur.execute(self.query, values)
if __name__ == '__main__':
    Linkedinparsing().main()

