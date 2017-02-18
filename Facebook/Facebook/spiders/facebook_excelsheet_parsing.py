import MySQLdb
import datetime
import os
import openpyxl as px
import glob
import md5
import json

class Fbparsing(object):

	def __init__(self):
		self.con = MySQLdb.connect(db='FACEBOOK',
		user='root', passwd='root',
		charset="utf8", host='localhost', use_unicode=True)
		self.cur = self.con.cursor()
		self.social_processing_path = '/root/Facebook/Facebook/spiders/excelfiles'
		self.query = 'insert into facebook_crawl(sk, url, content_type ,crawl_status, meta_data,created_at, modified_at) values(%s, %s, %s, %s, %s,now(), now()) on duplicate key update modified_at=now(), content_type=%s, crawl_status=0,meta_data=%s'

	def __del__(self):
		self.cur.close()
		self.con.close()

	def rep_spl(self, var):
        	"""To remove unwanted characters"""
	        var = str(var).replace('\r', '').replace('\n', '')\
        	.replace('\t', '').replace(u'\xa0', '').replace(u'\xc2\xa0', '').strip()
	        if 'none' in var.lower():
        	    var = ''
	        return var
	def main(self):
		files_list = glob.glob(self.social_processing_path+'/*.xlsx')
		for _file in files_list:
			ws_ = px.load_workbook(_file, use_iterators=True)
			sheet_list = ws_.get_sheet_names()
			for xl_ in sheet_list:
				sheet_ = ws_.get_sheet_by_name(name=xl_)
				row_check = 0
				email_address = facebook_profile = 0
				for row in sheet_.iter_rows():
			        	if row_check > 0:
				    		row_check += 1
				   		continue
				        counter = 0
				        for i in row:
						if not i.value: continue
						lower_head = i.value.lower()
					    	if 'email' in lower_head:
				    			email_address = counter
						elif 'facebook' in lower_head:
				    			facebook_profile = counter
						counter += 1
					break
			        m = 0
				final_indents = []
				_indents = {}
				for row in sheet_.iter_rows():
					temp_rows = []
					for i in row:
						temp_rows.append(i.value)
				    	final_indents.append(temp_rows)
				empty_dic = {}
				for row in final_indents:
				    	addr  = self.rep_spl(row[email_address])
				    	fbpro = self.rep_spl(row[facebook_profile])
					if not fbpro: continue
					if addr == 'Email': continue
					meta_date_from_browse = {}
					meta_date_from_browse.update({"email_address":addr, "mbasic_url":fbpro.replace('www','mbasic')})
					sk = md5.md5("%s%s"%(addr, fbpro)).hexdigest()
					values = (sk, fbpro, 'facebook', 0,json.dumps(meta_date_from_browse),'facebook', json.dumps(meta_date_from_browse))
					self.cur.execute(self.query, values)
				
if __name__ == '__main__':
    Fbparsing().main()

