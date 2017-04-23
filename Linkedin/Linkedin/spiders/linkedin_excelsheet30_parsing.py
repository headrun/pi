import MySQLdb
import datetime
import os
import openpyxl as px
import glob
import md5
import hashlib
import json
import re
import urllib

def xcode(text, encoding='utf8', mode='strict'):
	return text.encode(encoding, mode) if isinstance(text, unicode) else text

def md5(x):
	return hashlib.md5(xcode(x)).hexdigest()

def replacefun(text):
	text = text.replace('"','<>#<>').replace("'","<>##<>").replace(',','###').replace(u'\u2013','').strip()
	return text

def restore(text):
	text = text.replace('<>#<>','"').replace("<>##<>","'").replace('###',',')
	return text

def clean(text):
	if not text: return text
	value = text
	value = re.sub("&amp;", "&", value)
	value = re.sub("&lt;", "<", value)
	value = re.sub("&gt;", ">", value)
	value = re.sub("&quot;", '"', value)
	value = re.sub("&apos;", "'", value)

	return value

def normalize(text):
	return clean(compact(xcode(text)))

def compact( text, level=0):
	if text is None: return ''
	if level == 0:
	    text = text.replace("\n", " ")
	    text = text.replace("\r", " ")
	compacted = re.sub("\s\s(?m)", " ", text)
	if compacted != text:
	    compacted = compact(compacted, level+1)
	return compacted.strip()


class Linkedinparsing(object):

	def __init__(self):
		self.con = MySQLdb.connect(db='FACEBOOK',
		user='root', passwd='root',
		charset="utf8", host='localhost', use_unicode=True)
		self.cur = self.con.cursor()
		self.social_processing_path = '/root/Linkedin/Linkedin/spiders/excelfiles'
		self.query = 'insert into linkedin_crawl30(sk, url, clean_url, content_type ,crawl_status,flag, meta_data,created_at, modified_at) values(%s, %s, %s, %s, %s,%s,now(), now()) on duplicate key update modified_at=now(), content_type=%s, crawl_status=%s,meta_data=%s,sk =%s, flag=%s'

	def __del__(self):
		self.cur.close()
		self.con.close()

	def rep_spl(self, var):
        	"""To remove unwanted characters"""
	        var = str(var).replace('\r', '').replace('\n', '')\
        	.replace('\t', '').replace(u'\xa0', '').replace(u'\xc2\xa0', '').replace(u'\xc3\u0192\xc2\xb1',' ').strip()
	        if 'none' in var.lower():
        	    var = ''
	        return var
	def main(self):
		files_list = glob.glob(self.social_processing_path+'/*.xlsx')
	        if files_list:
			for _file in files_list:
				ws_ = px.load_workbook(_file, use_iterators=True)
				sheet_list = ws_.get_sheet_names()
				for xl_ in sheet_list:
					sheet_ = ws_.get_sheet_by_name(name=xl_)
					row_check = 0
					email_address = linkedin_profile = ida =  0
					for row in sheet_.iter_rows():
						if row_check > 0:
							row_check += 1
							continue
						counter = 0
						for i in row:
							if not i.value: continue
							lower_head = i.value.lower()
							"""if 'email' in lower_head:
								email_address = counter
							elif ('linkedin' in lower_head) or ('linkdin' in lower_head):
								linkedin_profile = counter"""
							if 'id' == lower_head.strip():
								ida = counter
							elif 'linkdin_url' in lower_head:
								linkedin_profile = counter
							counter += 1
						break
					m = 0
					final_indents = []
					_indents = {}
					for row in sheet_.iter_rows():
						temp_rows = []
						for i in row:
							temp_rows.append(i.value)
						final_indents.append(temp_rows)
					empty_dic = {}
					for row in final_indents:
						linkedin_profilef = normalize(row[linkedin_profile])
						idf = self.rep_spl(row[ida])
						#addr  = self.rep_spl(row[email_address])
						if 'linkdin_url' in linkedin_profilef: continue
						#fbpro = self.rep_spl(row[linkedin_profile])
						if not linkedin_profilef: continue
						meta_date_from_browse = {}
						meta_date_from_browse.update({"id":idf})
						meta_date_from_browse.update({"linkedin_url":linkedin_profilef})
						crawl_status = 0
						urlclear = linkedin_profilef
						linkedin_profilef = urllib.quote(linkedin_profilef)
						if 'linkedin.com' not in normalize(linkedin_profilef): crawl_status = 10
						linkedin_profilef = "%s%s%s"%('https://www.linkedin.com/cws/member/public_profile?public_profile_url=',linkedin_profilef, "&ffalse&original_referer=https%3A%2F%2Fdeveloper.linkedin.com%2Fetc%2Fdesigns%2Flinkedin%2Fkaty%2Fglobal%2Fclientlibs%2Fhtml%2Fsands%3Dmiddle-center&token=&isFramed=true&lang=en_US&_ts=1490264453590.3835&xd_origin_host=https%3A%2F%2Fdeveloper.linkedin.com")
                                                if 'linkedin.com' not in normalize(urlclear): crawl_status = 10
                                                if 'http:' in urlclear: urlclear = urlclear.replace('http:','https:')
                                                if 'id.www' in urlclear: urlclear = urlclear.replace('id.www','https://www')
                                                if 'www.linkedin.com' and 'https:' not in urlclear: urlclear = urlclear.replace('www.','https://www')
                                                if urlclear.startswith('linkedin.com'): urlclear = urlclear.replace('linkedin.com','https://www.linkedin.com')
                                                if 'https:' not in urlclear:
                                                        urlclear = re.sub('(\D+)\.linkedin.com','https://www.linkedin.com',urlclear)
                                                urlclear = re.sub('https://(.*?).linkedin.com/','https://www.linkedin.com/',urlclear)
                                                if urlclear.endswith('/en') or urlclear.endswith('/fr'): urlclear = urlclear[:-3]
                                                urlclear = urlclear.strip('"').strip().strip("'").strip().strip('/').strip()
                                                if not urlclear.startswith('https://www.linkedin.com') and crawl_status!=10: urlclear = ''.join(re.findall('.*(https://.*)', urlclear))
                                                if '/pub/' in urlclear:
                                                        cv = ''.join(filter(None,re.split('https://www.linkedin.com/pub/.*?/(.*)',urlclear))).split('/')[::-1]
                                                        cv[0] = cv[0].zfill(3)
                                                        cv[1] = cv[1].zfill(3)
                                                        if cv[-1] == '0': del cv[-1]
                                                        urlclear = ( '%s%s%s%s'%('https://www.linkedin.com/in/',''.join(re.findall('https://www.linkedin.com/pub/(.*?)/.*',urlclear)),'-',''.join(cv)))
						sk = md5("%s%s"%(normalize(idf),normalize( linkedin_profilef)))
						linkedin_profilef = linkedin_profilef.replace('pubwww.linkedin.comhttps:','').replace('"','')
						print linkedin_profilef
						values = (sk, linkedin_profilef, urlclear, 'linkedin', crawl_status,'', json.dumps(meta_date_from_browse),'linkedin', crawl_status, json.dumps(meta_date_from_browse),sk , '')
						self.cur.execute(self.query, values)
		else:
		    for prof_url in open('linkedin_file.py'):
			if prof_url != '\n':
				prof_url = prof_url.replace('\n','')
				sk = md5.md5(prof_url).hexdigest()
				meta_date_from_browse = {}
				values = (sk, prof_url, 'linkedin', 0,json.dumps(meta_date_from_browse),'linkedin', json.dumps(meta_date_from_browse))
				self.cur.execute(self.query, values)
if __name__ == '__main__':
    Linkedinparsing().main()

