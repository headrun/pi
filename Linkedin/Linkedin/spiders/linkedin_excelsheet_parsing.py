import MySQLdb
import datetime
import os
import openpyxl as px
import glob
import md5
import hashlib
import json
import re

def xcode(text, encoding='utf8', mode='strict'):
	return text.encode(encoding, mode) if isinstance(text, unicode) else text

def md5(x):
	return hashlib.md5(xcode(x)).hexdigest()

def replacefun(text):
	text = text.replace('"','<>#<>').replace("'","<>##<>").replace(',','###').replace(u'\u2013','').strip()
	return text

def restore(text):
	text = text.replace('<>#<>','"').replace("<>##<>","'").replace('###',',')
	return text

def clean(text):
	if not text: return text
	value = text
	value = re.sub("&amp;", "&", value)
	value = re.sub("&lt;", "<", value)
	value = re.sub("&gt;", ">", value)
	value = re.sub("&quot;", '"', value)
	value = re.sub("&apos;", "'", value)

	return value

def normalize(text):
	return clean(compact(xcode(text)))

def compact( text, level=0):
	if text is None: return ''
	if level == 0:
	    text = text.replace("\n", " ")
	    text = text.replace("\r", " ")
	compacted = re.sub("\s\s(?m)", " ", text)
	if compacted != text:
	    compacted = compact(compacted, level+1)
	return compacted.strip()


class Linkedinparsing(object):

	def __init__(self):
		self.con = MySQLdb.connect(db='FACEBOOK',
		user='root', passwd='root',
		charset="utf8", host='localhost', use_unicode=True)
		self.cur = self.con.cursor()
		self.social_processing_path = '/root/Linkedin/Linkedin/spiders/excelfiles'
		self.query = 'insert into linkedin_crawl(sk, url, content_type ,crawl_status, meta_data,created_at, modified_at) values(%s, %s, %s, %s, %s,now(), now()) on duplicate key update modified_at=now(), content_type=%s, crawl_status=0,meta_data=%s'

	def __del__(self):
		self.cur.close()
		self.con.close()

	def rep_spl(self, var):
        	"""To remove unwanted characters"""
	        var = str(var).replace('\r', '').replace('\n', '')\
        	.replace('\t', '').replace(u'\xa0', '').replace(u'\xc2\xa0', '').replace(u'\xc3\u0192\xc2\xb1',' ').strip()
	        if 'none' in var.lower():
        	    var = ''
	        return var
	def main(self):
		files_list = glob.glob(self.social_processing_path+'/*.xlsx')
	        if files_list:
			for _file in files_list:
				ws_ = px.load_workbook(_file, use_iterators=True)
				sheet_list = ws_.get_sheet_names()
				for xl_ in sheet_list:
					sheet_ = ws_.get_sheet_by_name(name=xl_)
					row_check = 0
					email_address = linkedin_profile = ida = firstname = lastname = 0
					for row in sheet_.iter_rows():
						if row_check > 0:
							row_check += 1
							continue
						counter = 0
						for i in row:
							if not i.value: continue
							lower_head = i.value.lower()
							"""if 'email' in lower_head:
								email_address = counter
							elif ('linkedin' in lower_head) or ('linkdin' in lower_head):
								linkedin_profile = counter"""
							if 'id' in lower_head:
								ida = counter
							elif 'firstname' in lower_head:
								firstname = counter
							elif 'lastname' in lower_head:
								lastname = counter
							elif 'linkdin_url' in lower_head:
								linkedin_profile = counter
							elif 'linkedin'  in lower_head:
								linkedin_profile = counter
							elif 'email' in lower_head:
								email_address = counter
							counter += 1
						break
					m = 0
					final_indents = []
					_indents = {}
					for row in sheet_.iter_rows():
						temp_rows = []
						for i in row:
							temp_rows.append(i.value)
						final_indents.append(temp_rows)
					empty_dic = {}
					counter_ = 0
					for row in final_indents:
						idf = self.rep_spl(row[ida])
						"""try: firstnamef = normalize(row[firstname])
						except: pass 
						try:
							lastnamef = row[lastname]
							if lastnamef: lastnamef= normalize(lastnamef.replace(u'\xc3\u0192\xc2\xb1',''))
						except: pass"""
						#email_addressf = normalize(row[email_address])
						linkedin_profilef = ''
						try: linkedin_profilef = normalize(row[linkedin_profile])
						except: pass
						if not linkedin_profilef: continue
						#if lastnamef == 'lastname': continue
						if linkedin_profilef.strip() == 'linkedin': continue
						meta_date_from_browse = {}
						meta_date_from_browse.update({"id":idf})
						#meta_date_from_browse.update({"firstname":firstnamef})
						#meta_date_from_browse.update({"lastname":lastnamef})
						meta_date_from_browse.update({"linkedin_url":linkedin_profilef})
						#meta_date_from_browse.update({"email_address":email_addressf})
						crawl_status = 0
						rows_ = []
						if linkedin_profilef.count('http') > 1: 
							inner_results = [i.replace('s://','https://').replace(', ht','').replace('LI Inmail Sent','').strip().strip(',') for i in filter(None,re.split('http*',linkedin_profilef))]
							for inr in inner_results:
								if (inr != 'https://de.linkedin.com/pub') and ('philippe-nieuwjaer') not in inr: rows_.append(inr)
						#else: rows_.append(linkedin_profilef)
						for rs in rows_:
							linkedin_profilef = rs
							if 'linkedin.com' not in normalize(linkedin_profilef): crawl_status = 10
							if 'http:' in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('http:','https:')
							if 'id.www' in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('id.www','https://www')
							if 'www.linkedin.com' and 'https:' not in linkedin_profilef: linkedin_profilef = linkedin_profilef.replace('www.','https://www')
							if linkedin_profilef.startswith('linkedin.com'): linkedin_profilef = linkedin_profilef.replace('linkedin.com','https://www.linkedin.com')
							if 'https:' not in linkedin_profilef:
								linkedin_profilef = re.sub('(\D+)\.linkedin.com','https://www.linkedin.com',linkedin_profilef) 
							linkedin_profilef = re.sub('https://(.*?).linkedin.com/','https://www.linkedin.com/',linkedin_profilef)
							if linkedin_profilef.endswith('/en') or linkedin_profilef.endswith('/fr'): linkedin_profilef = linkedin_profilef[:-3]
							linkedin_profilef = linkedin_profilef.strip('"').strip().strip("'").strip().strip('/').strip()
							if not linkedin_profilef.startswith('https://www.linkedin.com') and crawl_status!=10: linkedin_profilef = ''.join(re.findall('.*(https://.*)', linkedin_profilef))
							if '/pub/' in linkedin_profilef:
								cv = ''.join(filter(None,re.split('https://www.linkedin.com/pub/.*?/(.*)',linkedin_profilef))).split('/')[::-1]
								cv[0] = cv[0].zfill(3)
								cv[1] = cv[1].zfill(3)
								if cv[-1] == '0': del cv[-1]
								linkedin_profilef = ( '%s%s%s%s'%('https://www.linkedin.com/in/',''.join(re.findall('https://www.linkedin.com/pub/(.*?)/.*',linkedin_profilef)),'-',''.join(cv)))
							if 'linkedin.com' not in linkedin_profilef: continue
							counter_ += 1
							sk = md5("%s%s%s"%(normalize(idf),normalize( linkedin_profilef), str(counter_)))
							values = (sk, linkedin_profilef, 'linkedin', crawl_status, json.dumps(meta_date_from_browse),'linkedin', json.dumps(meta_date_from_browse))
							#print linkedin_profilef
							#print meta_date_from_browse
							#print '****************'
							self.cur.execute(self.query, values)
							print self.query%values
							print '*************************'
		else:
		    for prof_url in open('linkedin_file.py'):
			if prof_url != '\n':
				prof_url = prof_url.replace('\n','')
				sk = md5.md5(prof_url).hexdigest()
				meta_date_from_browse = {}
				values = (sk, prof_url, 'linkedin', 0,json.dumps(meta_date_from_browse),'linkedin', json.dumps(meta_date_from_browse))
				self.cur.execute(self.query, values)
if __name__ == '__main__':
    Linkedinparsing().main()

